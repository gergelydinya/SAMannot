import os
from PIL import Image
import cv2  
import time
import numpy as np
from label import Label_Handler
from sam_annotator import SAM_Annotator
from openpyxl import Workbook
import shutil
import re
from skimage.measure import label, regionprops
from skimage.morphology import skeletonize
from scipy import ndimage as ndi
import gc
class Annotator:
    def __init__(self):
        self.base_checkpoint_path = "./checkpoints/"
        self.init_sam2()
        self.sam_extract_dir = "sam_temp_dir"
        
        self.media_files = []
        self.curr_img_idx = -1
        self.current_block = 0
        self.idx_to_path = {}
        self.temp_dir = None
        self.video_name = ""
        self.media_path = "."
        
        self.overlay_img = None
        self.composite_mask = None
        self.masks = {}
        self.overlay_imgs = {}
        self.combined_masks = {}
        self.cache_images = True
        self.current_img_width = 0
        self.current_img_height = 0

        self.label_handler = Label_Handler()
        self.curr_label_idx = -1
        self.image_file_types=["jpg","JPG","jpeg","JPEG","png"]
        self.video_file_types=["mp4","avi","mov","mkv"]

        self.mode = "prompts"
        self.fn_modes = ["prompts","tracking","correction","export"]
        self.view_mode = "prompts"
        self.view_modes = ["original","prompts","overlay","masks"]
        self.session_name = "Session"
        
        self.tracking_results = {}
        self.extra_frame_path = {}
        self.extra_frame = {}
    
    # SESSION HANDLING
    
    def reset(self):
        self.media_files = []
        self.curr_img_idx = -1
        self.overlay_img = None
        self.composite_mask = None
        self.original_img = None
        self.overlay_imgs = {}
        self.combined_masks = {}
        self.masks = {}
        self.tracking_results = {}
        self.label_handler = Label_Handler()
        self.curr_label_idx = -1
        self.mode = "prompts"
        self.view_mode = "prompts"
        self.session_name = "Session"
        self.video_name = ""
        self.media_path = "."
        self.tracking_results = {}
        self.sam_handler.labels = []
        self.sam_handler.object_id_to_label_name = {}
        self.sam_handler.media_files = []
        self.sam_handler.propagation_blocks = {}
        self.current_block = 0
        self.extra_frame_path = {}
        self.extra_frame = {}
        self.extra_frame_masks = {}
        if self.sam_handler.inference_state:
            self.sam_handler.predictor.reset_state(self.sam_handler.inference_state)
    def load_from_dict(self,dict_representation):
        self.session_name = dict_representation["session_name"]
        self.view_mode = dict_representation["view_mode"]
        self.mode = dict_representation["mode"]
        self.media_files = dict_representation["media_files"]
        self.curr_img_idx = dict_representation["curr_img_idx"]
        self.overlay_img = dict_representation["overlay_img"]
        self.overlay_imgs = dict_representation["overlay_imgs"]
        self.combined_masks = dict_representation["combined_masks"]
        self.original_img = dict_representation["original_img"]
        self.composite_mask = dict_representation["composite_mask"]
        self.masks = dict_representation["masks"]
        self.idx_to_path = dict_representation["idx_to_path"]
        self.temp_dir = dict_representation["temp_dir"]
        self.current_img_width = dict_representation["current_img_width"]
        self.current_img_height = dict_representation["current_img_height"]
        self.label_handler = dict_representation["label_handler"]
        self.curr_label_idx = dict_representation["curr_label_idx"]
        self.tracking_results = dict_representation["tracking_results"]
        self.cache_images = dict_representation["cache_images"]
        self.media_path = dict_representation["media_path"]
        self.video_name = dict_representation["video_name"]
        self.sam_handler.labels = dict_representation["sam_handler_labels"]
        self.sam_handler.object_id_to_label_name = dict_representation["sam_handler_object_id_to_label_name"]
        self.sam_handler.media_files = dict_representation["sam_handler_media_files"]
        self.sam_handler.curr_img_idx = dict_representation["sam_handler_curr_img_idx"]
        self.sam_handler.current_stage = dict_representation["sam_handler_current_stage"]
        self.sam_handler.model_loaded  = dict_representation["sam_handler_model_loaded"]
        self.sam_handler.model_loading = dict_representation["sam_handler_model_loading"]
        self.sam_handler.propagation_blocks = dict_representation["sam_handler_propagation_blocks"]
        self.block_size = dict_representation["block_size"]
        self.current_block = dict_representation["current_block"]
        self.num_blocks = dict_representation["num_blocks"]
        self.curr_img_shape = dict_representation["curr_img_shape"]
        self.extra_frame_path = dict_representation["extra_frame_path"]
        self.extra_frame = dict_representation["extra_frame"]
        self.extra_frame_masks = dict_representation["extra_frame_masks"]
    def compress_to_dict(self):
        dict_representation = {
            "session_name": self.session_name,
            "view_mode": self.view_mode,
            "mode": self.mode,
            "media_files": self.media_files,
            "curr_img_idx": self.curr_img_idx,
            "overlay_img": self.overlay_img,
            "overlay_imgs": self.overlay_imgs,
            "combined_masks": self.combined_masks,
            "original_img": self.original_img,
            "composite_mask": self.composite_mask,
            "masks": self.masks,
            "idx_to_path": self.idx_to_path,
            "temp_dir": self.temp_dir,
            "current_img_width": self.current_img_width,
            "current_img_height": self.current_img_height,
            "label_handler": self.label_handler,
            "curr_label_idx": self.curr_label_idx,
            "tracking_results": self.tracking_results,
            "cache_images": self.cache_images,
            "media_path": self.media_path,
            "video_name": self.video_name,
            "block_size": self.block_size,
            "current_block": self.current_block,
            "num_blocks": self.num_blocks,
            "curr_img_shape": self.curr_img_shape,
            "extra_frame_path": self.extra_frame_path,
            "extra_frame": self.extra_frame,
            "extra_frame_masks": self.extra_frame_masks,
            "sam_handler_labels": self.sam_handler.labels,
            "sam_handler_object_id_to_label_name": self.sam_handler.object_id_to_label_name,
            "sam_handler_media_files": self.sam_handler.media_files,
            "sam_handler_curr_img_idx": self.sam_handler.curr_img_idx,
            "sam_handler_current_stage": self.sam_handler.current_stage,
            "sam_handler_model_loaded": self.sam_handler.model_loaded,
            "sam_handler_model_loading": self.sam_handler.model_loading,
            "sam_handler_propagation_blocks": self.sam_handler.propagation_blocks,
        }
        return dict_representation

    # OTHERS

    def set_session_name(self,session_name):
        self.session_name = session_name
    def get_session_name(self):
        return self.session_name
    def set_cache_images(self,value):
        if value == False:
            self.overlay_imgs = {}
            self.combined_masks = {}
            gc.collect()
        self.cache_images = value
    def check_label_existence(self, label_name):
        for label in self.sam_handler.labels:
            if label.name == label_name:
                return True
        return False
    def get_current_block(self):
        return self.current_block
    def set_num_blocks(self,num_blocks):
        self.num_blocks = num_blocks
        for i in range(self.num_blocks):
            self.sam_handler.propagation_blocks[i] = {}
    def set_current_block(self,block):
        if block < 0:
            block = 0
        self.current_block = block
        self.sam_handler.current_block = self.current_block
    def has_labels(self):
        return len(self.sam_handler.labels) > 0
    def has_frames(self):
        return len(self.media_files) > 0
    def has_prompts(self, block_idx):
        for label in self.get_labels():
            if len(label.pts[block_idx]) > 0 or len(label.boxes[block_idx]):
                return True
        return False
    def get_label_idx(self, label_name):
        for idx in range(len(self.sam_handler.labels)):
            if self.sam_handler.labels[idx].name == label_name:
                return idx
        return -1
    def get_current_label(self):
        if self.curr_label_idx < 0:
            return None
        return self.sam_handler.labels[self.curr_label_idx]
    def set_current_label(self,idx):
        self.curr_label_idx = idx
    def set_label_name(self, idx, name):
        old_name = self.sam_handler.labels[idx].name
        new_name = name
        for masks in self.masks.values():
            if old_name in masks:
                masks[new_name] = masks.pop(old_name)
        self.sam_handler.labels[idx].name = name
    def get_labels(self):
        return self.sam_handler.labels
    def get_current_label_idx(self):
        return self.curr_label_idx
    def get_current_img_idx(self):
        return self.curr_img_idx
    def get_number_of_images(self):
        return len(self.media_files)
    
    # LABEL AND PROMPT HANDLING
    
    def add_label(self,label_name):
        new_label = self.label_handler.create_new_label(label_name)
        for i in range(self.num_blocks):
            new_label.pts[i] = []
        for i in range(self.num_blocks):
            new_label.boxes[i] = []
        for i in range(self.num_blocks):
            new_label.prop_frames[i] = set()
        self.sam_handler.labels.append(new_label) 
        self.curr_label_idx = len(self.sam_handler.labels) - 1
        return new_label.col
    def add_feature_to_current_label(self,feature_name):
        current_label = self.sam_handler.labels[self.curr_label_idx]
        current_label.features.append([(feature_name, self.curr_img_idx + self.current_block * self.block_size)])
    def delete_feature_from_current_label(self,feature_idx):
        current_label = self.sam_handler.labels[self.curr_label_idx]
        current_label.features.pop(feature_idx)
    def delete_selected_point(self,point_idx):
        current_label = self.sam_handler.labels[self.curr_label_idx]
        if point_idx < len(current_label.pts[self.current_block]):
            current_label.remove_pt(point_idx,self.current_block)
    def delete_selected_box(self,box_idx):
        current_label = self.sam_handler.labels[self.curr_label_idx]
        if box_idx < len(current_label.boxes[self.current_block]):
            current_label.remove_box(box_idx,self.current_block)
    
    # MODEL HANDLING
    
    def init_sam2(self):
        self.sam_handler = SAM_Annotator(model_type="vit_h",model_cfg_path="sam2.1_hiera_l.yaml",ckpt_path=os.path.join(os.path.dirname(os.path.abspath(__file__)), self.base_checkpoint_path+"sam2.1_hiera_large.pt"))
    def can_initialize_model(self):
        return os.path.exists(self.sam_handler.ckpt_path)
    def get_model_type(self):
        return self.sam_handler.model_type
    def model_status(self):
        return self.sam_handler.is_model_loaded()
    def load_model(self,status_callback):
        if not self.can_initialize_model():
            return False
        return self.sam_handler.load_model(status_callback)
    
    # PROPAGATION

    def initialize_tracking(self,update_progress_callback):
        self.sam_handler.media_files = self.media_files
        self.sam_handler.curr_img_idx = self.curr_img_idx
        self.mode = "tracking"
        success = self.sam_handler.init_inference_state(self.sam_extract_dir, status_callback=update_progress_callback)
        if not success:
            self.mode = "prompts"
        return success
    def apply_masks(self,update_progress_callback):
        total_steps = len(list(self.tracking_results.keys())) 
        current_step = 0
        for frame_idx, frame_masks in self.tracking_results.items():
            if frame_idx < len(self.media_files):
                frame_path = self.media_files[frame_idx]
                self.idx_to_path[frame_idx] = frame_path
                self.masks[frame_path] = {}
                try:
                    for label_name, mask in frame_masks.items():
                        label_color = None
                        for label in self.sam_handler.labels:
                            if label.name == label_name:
                                label_color = self.hex_to_rgb(label.col)
                                break
                        if label_color is None:
                            label_color = (255, 255, 255)
                        self.masks[frame_path][label_name]=PackedMasks(mask.squeeze().astype(bool))
                    progress_percent = (current_step / total_steps) * 100
                    update_progress_callback(f"Applying masks: {current_step + 1}/{total_steps}", progress_percent)
                    current_step += 1
                    if self.cache_images:
                        self.combined_masks[frame_path] = self.create_combined_mask(frame_path,overwrite=True)
                        self.overlay_imgs[frame_path] = self.create_overlay_img(frame_path,overwrite=True)
                except Exception as e:
                    print(f"Error generating mask for frame {frame_idx}: {str(e)}")
        self.tracking_results = {}
        self.composite_mask = self.create_combined_mask(overwrite=True)
        self.overlay_img = self.create_overlay_img(overwrite=True)
        if self.cache_images:
            self.combined_masks[self.media_files[self.curr_img_idx]] = self.composite_mask
            self.overlay_imgs[self.media_files[self.curr_img_idx]] = self.overlay_img
        self.mode = "correction"
        self.view_mode = "overlay"
    def has_propagation_block(self,idx):
        return (idx in self.sam_handler.propagation_blocks[self.current_block])
    def add_propagation_block(self,idx):
        self.sam_handler.propagation_blocks[self.current_block][idx] = 1
    def remove_propagation_block(self,idx):
        del self.sam_handler.propagation_blocks[self.current_block][idx]
    def generate_mask(self):
        if not self.sam_handler.is_model_loaded():
            return False, []
        if self.curr_label_idx < 0 or len(self.media_files) < 0:
            return False, []
        labels_with_points = [label for label in self.sam_handler.labels if len(label.pts[self.current_block]) > 0]
        if not labels_with_points:
            return False, []
        self.sam_handler.media_files = self.media_files
        self.sam_handler.curr_img_idx = self.curr_img_idx
        self.tracking_results = self.sam_handler.generate_mask_for_frame(self.curr_img_idx,0)
        if self.tracking_results is None:
            self.mode = "prompts"
            return False, []
        else:
            for k, v in self.tracking_results.items():
                for label,_ in v.items():
                    self.sam_handler.labels[self.get_label_idx(label)].prop_frames[self.current_block].add(k)
        return True, list(self.tracking_results.keys())
    def propagate(self,flag,update_progress_callback):
        if flag == 0:
            self.tracking_results = self.sam_handler.propagate_to_all(
                current_frame_idx=self.curr_img_idx,
                start_frame_idx=None,
                end_frame_idx=None,
                progress_callback=update_progress_callback)
        elif flag == 1:
            self.tracking_results = self.sam_handler.propagate(1,
                start_frame_idx=self.curr_img_idx,
                end_frame_idx=None,
                progress_callback=update_progress_callback)
        else:
            self.tracking_results = self.sam_handler.propagate(-1,
                start_frame_idx=self.curr_img_idx,
                end_frame_idx=None,
                progress_callback=update_progress_callback)
        if self.tracking_results is None:
            self.mode = "prompts"
            return False, []
        else:
            if self.block_size in self.tracking_results:
                self.extra_frame[self.current_block] = self.tracking_results[self.block_size]
                self.tracking_results.pop(self.block_size)
                self.extra_frame_masks[self.current_block] = {}
                for label_name, mask in self.extra_frame[self.current_block].items():
                    label_color = None
                    for label in self.sam_handler.labels:
                        if label.name == label_name:
                            label_color = self.hex_to_rgb(label.col)
                            break
                    if label_color is None:
                        label_color = (255, 255, 255)
                    self.extra_frame_masks[self.current_block][label_name]=PackedMasks(mask.squeeze().astype(bool))
            for k, v in self.tracking_results.items():
                for label,_ in v.items():
                    self.sam_handler.labels[self.get_label_idx(label)].prop_frames[self.current_block].add(k)
        return True, list(self.tracking_results.keys())
    
    # AUTO-PROMPTING ~ DONE
    
    def keep_largest_connected_region(self,mask):
        labeled = label(mask > 0)
        if labeled.max() == 0:
            return np.zeros_like(mask)
        largest_region = max(regionprops(labeled), key=lambda r: r.area)
        return (labeled == largest_region.label).astype(np.uint8) * 255
    def skeletonize_single_mask(self,mask):
        if mask is None:
            return None
        mask = self.keep_largest_connected_region(mask)
        mask = mask.astype(np.uint8)
        skeleton = skeletonize(mask).astype(np.uint8)
        return skeleton
    def generate_point_prompts(self,mask):
        skeleton = self.skeletonize_single_mask(mask)
        neighbourhood_kernel = np.array([[1, 1, 1],
                                        [1, 0, 1],
                                        [1, 1, 1]], dtype=np.uint8)
        neighbor_count = ndi.convolve(skeleton.astype(np.uint8),neighbourhood_kernel,mode='constant', cval=0)
        endpoints = skeleton & (neighbor_count == 1)
        endpoints = np.vstack(np.nonzero(endpoints)).T
        junctions = skeleton & (neighbor_count >= 3)
        structure = np.ones((3, 3), dtype=np.uint8)
        labels, n_labels = ndi.label(junctions, structure=structure)
        if n_labels > 0:
            junctions = np.asarray(ndi.center_of_mass(junctions, labels, index=range(1, n_labels + 1)),dtype=np.int32)
            return np.vstack([endpoints,junctions])
        else:
            junctions = np.asarray([])
            return endpoints
    
    # VIEW MODE
    
    def get_next_view_mode(self):
        if self.view_mode == "original":
            return "prompts"
        elif self.view_mode == "prompts":
            if hasattr(self, 'overlay_img') and self.media_files[self.curr_img_idx] in self.masks:
                return "overlay"
            elif hasattr(self, 'masks') and self.media_files[self.curr_img_idx] in self.masks:
                return "masks"
            else:
                return "original"
        elif self.view_mode == "overlay":
            if hasattr(self, 'masks') and self.media_files[self.curr_img_idx] in self.masks:
                return "masks"
            else:
                return "original"
        else:
            return "original"
    def allowed_view_mode(self,new_view_mode):
        if new_view_mode == "original":
            return True
        elif new_view_mode == "prompts":
            return True
        elif new_view_mode == "overlay":
            return (hasattr(self, 'overlay_img') and self.media_files[self.curr_img_idx] in self.masks)
        else:
            return (hasattr(self, 'masks') and self.media_files[self.curr_img_idx] in self.masks)
    def set_view_mode(self, vm_idx):
        if vm_idx < 0 or vm_idx > 3:
            return 
        new_view_mode = self.view_modes[vm_idx]
        if self.allowed_view_mode(new_view_mode):
            self.view_mode = new_view_mode
    def toggle_view_mode(self):
        self.view_mode = self.get_next_view_mode()
    def reset_view_mode(self):
        self.mode = "prompts"
        self.view_mode = "original"
    def get_view_mode(self):
        return self.view_mode
    def get_mode(self):
        return self.mode
    
    # ANNOTATION EXPORT
    
    def bbox_from_mask(self,mask):
        if not np.any(mask):
            return None, None
        r = np.any(mask, axis=1)
        c = np.any(mask, axis=0)
        rmin,rmax = np.where(r)[0][[0,-1]]
        cmin,cmax = np.where(c)[0][[0,-1]]
        center = np.argwhere(mask).mean(axis=0)
        return [rmin,rmax,cmin,cmax],tuple(center)
    def export_manual_data(self):
        wb = Workbook()
        ws = wb.active
        session_id = self.session_name.split("_")
        frame_indices = list(self.idx_to_path.keys())
        frame_indices.sort()
        row = 1
        col = ord("A")
        ws[chr(col)+str(row)] = "frame_id"
        col = col+1
        for section in session_id:
            ws[chr(col)+str(row)] = "session_id"
            col = col+1
        ws[chr(col)+str(row)] = "label"
        col = col+1
        ws[chr(col)+str(row)] = "features"
        col = col+1
        ws[chr(col)+str(row)] = "clicks"
        col = col+1
        ws[chr(col)+str(row)] = "averaged click x"
        col = col+1
        ws[chr(col)+str(row)] = "averaged click y"
        col = col+1
        ws[chr(col)+str(row)] = "boxes"
        col = col+1
        row += 1
        label_dict = {k.name:k for k in self.sam_handler.labels}
        frame_paths = list(self.masks.keys())
        frame_paths.sort()
        idx = 0
        points = {}
        boxes = {}
        avg_points = {}
        for label in self.sam_handler.labels:
            for k,v in label.pts.items():
                for pt in v:
                    if (pt.idx + self.block_size * k) not in points:
                        points[pt.idx + self.block_size * k] = {}
                        avg_points[pt.idx + self.block_size * k] = {}
                    if label.name not in points[pt.idx + self.block_size * k]:
                        points[pt.idx + self.block_size * k][label.name] = []
                        avg_points[pt.idx + self.block_size * k][label.name] = [0,0,0]
                    points[pt.idx + self.block_size * k][label.name].append((pt.x,pt.y,pt.pt_type))
                    avg_points[pt.idx + self.block_size * k][label.name][0] += pt.x
                    avg_points[pt.idx + self.block_size * k][label.name][1] += pt.y
                    avg_points[pt.idx + self.block_size * k][label.name][2] += pt.pt_type
            for k,v in label.boxes.items():
                for box in v:
                    if (box.idx + self.block_size * k) not in boxes:
                        boxes[box.idx + self.block_size * k] = {}
                    if label.name not in boxes[box.idx + self.block_size * k]:
                        boxes[box.idx + self.block_size * k][label.name] = []              
                    boxes[box.idx + self.block_size * k][label.name].append((box.fx,box.fy,box.x,box.y,box.pt_type))
        point_indices = list(set(list(points.keys()) + list(boxes.keys())))
        point_indices.sort()
        for idx in point_indices:
            if idx in points:
                for k,v in points[idx].items():
                    col = ord("A")
                    ws[chr(col)+str(row)] = idx
                    col = col+1
                    for section in session_id:
                        ws[chr(col)+str(row)] = section
                        col = col+1
                    ws[chr(col)+str(row)] = k
                    col = col+1
                    ws[chr(col)+str(row)] = ','.join(str(x) for x in label_dict[k].features)
                    col = col+1
                    ws[chr(col)+str(row)] = ','.join(str(x) for x in v)
                    col = col+1
                    ws[chr(col)+str(row)] = avg_points[idx][k][0] / avg_points[idx][k][2]
                    col = col+1
                    ws[chr(col)+str(row)] = avg_points[idx][k][1] / avg_points[idx][k][2]
                    col = col+1
                    ws[chr(col)+str(row)] = ""
                    col = col+1
                    row += 1
            if idx in boxes:
                for k,v in boxes[idx].items():
                    col = ord("A")
                    ws[chr(col)+str(row)] = idx
                    col = col+1
                    for section in session_id:
                        ws[chr(col)+str(row)] = section
                        col = col+1
                    ws[chr(col)+str(row)] = k
                    col = col+1
                    ws[chr(col)+str(row)] = ','.join(str(x) for x in label_dict[k].features)
                    col = col+1
                    ws[chr(col)+str(row)] = ""
                    col = col+1
                    ws[chr(col)+str(row)] = ""
                    col = col+1
                    ws[chr(col)+str(row)] = ""
                    col = col+1
                    ws[chr(col)+str(row)] = ','.join(str(x) for x in v)
                    col = col+1
                    row += 1
        wb.save(os.path.join(os.getcwd(),"export",self.session_name,"manual_data.xlsx"))
    def export_interpolated_data(self,update_progress_callback):
        wb = Workbook()
        ws = wb.active
        session_id = self.session_name.split("_")
        frame_indices = list(self.idx_to_path.keys())
        frame_indices.sort()
        row = 1
        col = ord("A")
        ws[chr(col)+str(row)] = "frame_id"
        col = col+1
        for section in session_id:
            ws[chr(col)+str(row)] = "session_id"
            col = col+1
        ws[chr(col)+str(row)] = "label"
        col = col+1
        ws[chr(col)+str(row)] = "features"
        col = col+1
        ws[chr(col)+str(row)] = "bounding box"
        col = col+1
        ws[chr(col)+str(row)] = "center-x"
        col = col+1
        ws[chr(col)+str(row)] = "center-y"
        col = col+1
        row += 1
        label_dict = {k.name:k for k in self.sam_handler.labels}
        frame_paths = list(self.masks.keys())
        frame_paths.sort()
        for idx, path in enumerate(frame_paths):
            for k,v in self.masks[path].items():
                col = ord("A")
                ws[chr(col)+str(row)] = idx
                col = col+1
                for section in session_id:
                    ws[chr(col)+str(row)] = section
                    col = col+1
                bbox,center = self.bbox_from_mask(v)
                ws[chr(col)+str(row)] = k
                col = col+1
                feature_list = ""
                for i, feature in enumerate(label_dict[k].features):
                    feature_list += label_dict[k].find_most_recent_feature(i, idx)
                    feature_list += ","
                ws[chr(col)+str(row)] = feature_list
                col = col+1
                if bbox is not None:
                    ws[chr(col)+str(row)] = ','.join(str(x) for x in bbox)
                    col = col+1
                    ws[chr(col)+str(row)] = int(center[0])
                    col = col+1
                    ws[chr(col)+str(row)] = int(center[1])
                    col = col+1
                else:
                    ws[chr(col)+str(row)] = '-'
                    col = col+1
                    ws[chr(col)+str(row)] = '-'
                    col = col+1
                    ws[chr(col)+str(row)] = '-'
                    col = col+1
                row += 1
            update_progress_callback("interpolated",idx,len(frame_paths))
        wb.save(os.path.join(os.getcwd(),"export",self.session_name,"interpolated_data.xlsx"))
    def export_overlays(self,update_progress_callback):
        current_step = 0
        total_steps = len(list(self.masks.keys()))
        for k,v in self.masks.items():
            cv2.imwrite(os.path.join(os.getcwd(),"export",self.session_name,"overlays",os.path.basename(k)),cv2.cvtColor(np.array(self.create_overlay_img(k)),cv2.COLOR_BGR2RGB))
            update_progress_callback("overlay",current_step,total_steps)
            current_step += 1
    def export_masks(self,update_progress_callback):
        current_step = 0
        total_steps = len(list(self.masks.keys()))
        label_to_id = {label.name: i for i, label in enumerate(self.sam_handler.labels)}
        for k in self.masks.keys():
            combined_mask = cv2.cvtColor(np.array(self.create_combined_mask(k)),cv2.COLOR_BGR2RGB)
            h, w = combined_mask.shape[:2]
            cv2.imwrite(os.path.join(os.getcwd(),"export",self.session_name,"masks",os.path.basename(k).split('.')[0] + ".png"), combined_mask)
            with open(os.path.join(os.getcwd(),"export",self.session_name,"labels",os.path.basename(k).split('.')[0] + ".txt"), "w") as f:
                for label in self.sam_handler.labels:
                    if label.name not in self.masks[k]:
                        continue
                    class_id = label_to_id[label.name]
                    try:
                        contours, _ = cv2.findContours(self.masks[k][label.name].astype(np.uint8) * 255, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                        for contour in contours:
                            if len(contour) < 3:
                                continue
                            contour = contour.squeeze().astype(float)
                            contour[:, 0] /= w
                            contour[:, 1] /= h
                            coords = " ".join([f"{x:.6f} {y:.6f}" for x, y in contour])
                            f.write(f"{class_id} {coords}\n")
                    except Exception as e:
                        print(f"Error in exporting masks: {str(e)}")
            img_rgb = cv2.imread(k)
            cv2.imwrite(os.path.join(os.getcwd(),"export",self.session_name,"frames",os.path.basename(k).split('.')[0]+".png"), img_rgb)
            update_progress_callback("mask",current_step,total_steps)
            current_step += 1
    def export_label_data(self):
        wb = Workbook()
        ws = wb.active
        row = 1
        col = ord("A")
        ws[chr(col)+str(row)] = "label"
        col = col+1
        ws[chr(col)+str(row)] = "color"
        col = col+1
        row += 1
        for label in self.sam_handler.labels:
            col = ord("A")
            ws[chr(col)+str(row)] = label.name
            col = col+1
            ws[chr(col)+str(row)] = str(self.hex_to_rgb(label.col))
            col = col+1
            row += 1
        wb.save(os.path.join(os.getcwd(),"export",self.session_name,"label_data.xlsx"))
    def init_export(self):
        shutil.rmtree(os.path.join(os.getcwd(),"export",self.session_name), ignore_errors=True)
        os.makedirs(os.path.join(os.getcwd(),"export",self.session_name),exist_ok=True)
        os.makedirs(os.path.join(os.getcwd(),"export",self.session_name,"overlays"),exist_ok=True)
        os.makedirs(os.path.join(os.getcwd(),"export",self.session_name,"masks"),exist_ok=True)
        os.makedirs(os.path.join(os.getcwd(),"export", self.session_name,"labels"), exist_ok=True)
        os.makedirs(os.path.join(os.getcwd(),"export", self.session_name,"frames"), exist_ok=True)
        
    # MEDIA HANDLING
    
    def reset_media(self):
        self.media_files = []
        self.curr_img_idx = -1
        self.overlay_img = None
        self.composite_mask = None
        self.original_img = None
        self.overlay_imgs = {}
        self.combined_masks = {}
        self.mode = "prompts"
        self.view_mode = "prompts"
        self.tracking_results = {}
        self.sam_handler.object_id_to_label_name = {}
        self.sam_handler.media_files = []
        self.curr_img_shape = (0,0,0)
        if self.sam_handler.inference_state:
            self.sam_handler.predictor.reset_state(self.sam_handler.inference_state)
    def clear_temp_dir(self):
        self.temp_dir = "temp_dir"
        shutil.rmtree(self.temp_dir, ignore_errors=True)
        os.makedirs(self.temp_dir, exist_ok=True)
    def load_main_folder_unified(self, file_path, block_size):
        self.block_size = block_size
        if file_path is not None and len(file_path) > 0:
            extension = file_path.split(".")[-1]
            if extension in self.image_file_types:
                self.process_image_folder(os.path.dirname(file_path),self.current_block * block_size,(self.current_block + 1) * block_size)
                return 0
            elif os.path.isdir(file_path):
                self.process_image_folder(file_path,self.current_block * block_size,(self.current_block + 1) * block_size)
                return 0
            elif extension in self.video_file_types:
                self.process_video_file(file_path)
                return 1
            else:
                print(f"Invalid file format!")
                return -1
    def process_image_folder(self, folder_path, start_frame, end_frame):
        self.media_path = folder_path
        self.video_name = ""
        try:
            self.extract_dir = os.path.join(self.temp_dir)
            all_files = os.listdir(folder_path)
            self.media_files = [os.path.join(folder_path, file) for file in all_files if file.split(".")[-1] in self.image_file_types]
            def numeric_key( p):
                name = os.path.splitext(os.path.basename(p))[0]
                match = re.search(r'\d+', name)
                if match:
                    return int(match.group())
                return float('inf') 
            self.media_files.sort(key=numeric_key)
            end_frame = np.amin([end_frame,len(self.media_files)])
            self.media_files = self.media_files[start_frame:end_frame]
            shutil.rmtree(self.sam_extract_dir, ignore_errors=True)
            os.makedirs(self.sam_extract_dir, exist_ok=True)
            for file_name in self.media_files:
                shutil.copy(file_name,os.path.join(self.extract_dir,os.path.basename(file_name)))
                shutil.copy(file_name,os.path.join(self.sam_extract_dir,os.path.basename(file_name)))
            self.curr_img_idx = -1
        except Exception as e:
            print(f"Error loading folders: {str(e)}")
    def process_video_file(self, video_path):
        try:
            self.media_path = os.path.dirname(video_path)
            cap = cv2.VideoCapture(video_path)
            if not cap.isOpened():
                return
            cap.release()
            self.video_name = video_path
            self.extract_dir = self.temp_dir
            self.extracted_frames = []
        except Exception as e:
            print(f"Error processing video: {str(e)}")
    def get_loaded_frame_count(self):
        return len(self.media_files)
    def get_frame_count(self,video_path):
        total_frames = 0
        try:
            cap = cv2.VideoCapture(video_path)
            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            cap.release()
        except Exception as e:
            print(f"Error: {str(e)}")
        return total_frames
    def get_frame_count_dir(self,folder_path):
        temp_media_files = [os.path.join(folder_path, file) for file in os.listdir(folder_path) if file.split(".")[-1] in self.image_file_types]
        return len(temp_media_files)
    def extract_frames(self,start_frame,end_frame,interval,progress_var):
        cap = cv2.VideoCapture(self.video_name)
        frame_count = 0
        saved_count = 0
        for i in range(0,start_frame):
            ret, frame = cap.read()
            if not ret:
                break
        self.extracted_frames = []
        shutil.rmtree(self.sam_extract_dir, ignore_errors=True)
        os.makedirs(self.sam_extract_dir, exist_ok=True)
        for i in range(start_frame,end_frame):
            ret, frame = cap.read()
            if not ret:
                break
            if frame_count % interval == 0:
                frame_path = os.path.join(self.extract_dir, f"{i:06d}.jpg")
                self.extracted_frames.append(frame_path)
                cv2.imwrite(frame_path, frame)
                cv2.imwrite(os.path.join(self.sam_extract_dir, f"{i:06d}.jpg"), frame)
                saved_count += 1
            frame_count += 1
            progress = (frame_count / (end_frame - start_frame)) * 100
            progress_var.set(progress)
        ret, frame = cap.read()
        if ret:
            frame_path = os.path.join(self.extract_dir, f"{(i+1):06d}.jpg")
            self.extra_frame_path[self.current_block] = frame_path
            cv2.imwrite(os.path.join(self.sam_extract_dir, f"{(i+1):06d}.jpg"), frame)
        else:
            self.extra_frame_path[self.current_block] = None
        self.media_files = self.extracted_frames
        self.curr_img_idx = -1
        cap.release()
        
    # RENDER FRAME
    
    def hex_to_rgb(self, hex_color):
        hex_color = hex_color.lstrip('#')
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    def create_overlay_img(self,path = None,overwrite = False):
        alpha = 0.5
        if path is None:
            file_path = self.media_files[self.curr_img_idx]
        else:
            file_path = path
        if self.cache_images and path in self.overlay_imgs and not overwrite:
            return self.overlay_imgs[file_path]
        img_rgb = cv2.cvtColor(cv2.imread(file_path), cv2.COLOR_BGR2RGB)
        self.curr_img_shape = img_rgb.shape
        blended = cv2.addWeighted(img_rgb, 1, self.create_combined_mask(path,overwrite=overwrite), alpha, 0)
        return blended
    def create_combined_mask(self,path = None, overwrite = False):
        if path is None:
            path = self.media_files[self.curr_img_idx]
        if self.cache_images and (path in self.combined_masks) and (not overwrite):
            return self.combined_masks[path]
        mask_overlay = np.zeros(self.curr_img_shape,dtype=np.uint8)
        for label_name, mask in self.masks[path].items():
            label_color = None
            for label in self.sam_handler.labels:
                if label.name == label_name:
                    label_color = self.hex_to_rgb(label.col)
                    break
            if label_color is None:
                label_color = (255, 255, 255)
            mask_overlay[self.masks[path][label_name]] = label_color
        return mask_overlay
    def get_img_to_resize(self):
        if self.view_mode == "overlay":
            if not hasattr(self, 'masks'):
                return None
            return self.create_overlay_img()
        elif self.view_mode == "masks":
            if not hasattr(self, 'masks'):
                return None
            return self.create_combined_mask()
        else:
            if not hasattr(self, 'original_img'):
                return None
            if self.curr_img_idx == -1:
                return None
            original_img = cv2.cvtColor(cv2.imread(self.media_files[self.curr_img_idx]), cv2.COLOR_BGR2RGB)
            self.curr_img_shape = original_img.shape
            return original_img
        
    # ADD PROMPT
    
    def add_point_prompt_to_label(self,x, y, pt_type,frame_idx, label_idx):
        self.sam_handler.labels[label_idx].add_pt(x, y, pt_type,frame_idx,self.current_block)
    def add_box_prompt_to_label(self,fx,fy, x, y, pt_type,frame_idx, label_idx):
        self.sam_handler.labels[label_idx].add_box(fx, fy, x, y, pt_type,frame_idx,self.current_block)
    def add_point_prompt_to_current_label(self,x, y, pt_type,frame_idx):
        self.sam_handler.labels[self.curr_label_idx].add_pt(x, y, pt_type,frame_idx,self.current_block)
    def add_box_prompt_to_current_label(self,fx,fy, x, y, pt_type,frame_idx):
        self.sam_handler.labels[self.curr_label_idx].add_box(fx, fy, x, y, pt_type,frame_idx,self.current_block)
        
    # SWITCH IMG
    
    def load_img(self, image_path):
        self.original_img = Image.open(image_path)
    def set_img(self,idx):
        if not self.media_files:
            return
        self.curr_img_idx = idx
        self.refresh_img_data()
    def next_img(self):
        if not self.media_files:
            return
        if self.curr_img_idx < len(self.media_files) - 1:
            self.curr_img_idx += 1
            self.refresh_img_data()
    def prev_img(self):
        if not self.media_files:
            return
        if self.curr_img_idx > 0:
            self.curr_img_idx -= 1
            self.refresh_img_data()
    def refresh_img_data(self):
        if self.view_mode == "original":
            self.load_img(self.media_files[self.curr_img_idx])
        elif self.view_mode == "prompts":
            self.load_img(self.media_files[self.curr_img_idx])
        elif self.view_mode == "overlay":
            if self.media_files[self.curr_img_idx] not in self.masks:
                self.load_img(self.media_files[self.curr_img_idx])
                self.view_mode = "original"
            else:
                overlay_img = self.create_overlay_img()
                if not isinstance(overlay_img,Image.Image):
                    self.overlay_img = Image.fromarray(overlay_img)
                else:
                    self.overlay_img = overlay_img
        else:
            if self.media_files[self.curr_img_idx] not in self.masks:
                self.load_img(self.media_files[self.curr_img_idx])
                self.view_mode = "original"
            else:
                composite_mask = self.create_combined_mask()
                if not isinstance(composite_mask,Image.Image):
                    self.composite_mask = Image.fromarray(composite_mask)
                else:
                    self.composite_mask = composite_mask
class PackedMasks:
    def __init__(self, arr):
        arr = np.asarray(arr, dtype=bool)
        self.data = np.packbits(arr, axis=-1)
        self.h, self.w = arr.shape
    def _unpack(self):
        return np.unpackbits(self.data, axis=-1)[:, :self.w].astype(bool)
    def get(self):
        return self._unpack()
    def set(self, arr):
        arr = np.asarray(arr, dtype=bool)
        self.data = np.packbits(arr, axis=-1)
        self.h, self.w = arr.shape
    def __array__(self, dtype=None):
        arr = self._unpack()
        return arr.astype(dtype) if dtype is not None else arr
    def __getitem__(self, key):
        return self._unpack()[key]
    def __setitem__(self, key, value):
        arr = self._unpack()
        arr[key] = value
        self.set(arr)
    def _binary_op(self, other, op):
        return op(self._unpack(), other)
    def astype(self, dtype, **kwargs):
        return self._unpack().astype(dtype, **kwargs)
    def __gt__(self, other):
        return self._binary_op(other, np.greater)

    def __ge__(self, other):
        return self._binary_op(other, np.greater_equal)

    def __lt__(self, other):
        return self._binary_op(other, np.less)

    def __le__(self, other):
        return self._binary_op(other, np.less_equal)

    def __eq__(self, other):
        return self._binary_op(other, np.equal)

    def __ne__(self, other):
        return self._binary_op(other, np.not_equal)